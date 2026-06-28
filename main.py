import sys
import math
import json
import os
import ctypes
from PyQt6.QtGui import QIcon
from PyQt6.QtWidgets import QApplication, QMessageBox, QTableWidgetItem, QFileDialog
from openpyxl import Workbook, load_workbook
from PyQt6.QtCore import QThread, pyqtSignal, Qt, QSettings

from data_models import SystemState, BusData, LineData, CableConfig
from ui_main import MainWindowUI
from engine_sep import PowerSystemEngine
from plot_utils import populate_table

# Função para obter o caminho absoluto dos recursos (imagens) de forma compatível com PyInstaller
def resource_path(relative_path):
    """ Obtém o caminho absoluto para o recurso, funciona para dev e para PyInstaller """
    try:
        # PyInstaller cria uma pasta temporária e armazena o caminho em _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)

# Função para converter string em float de forma segura
def safe_float(val_str):
    if isinstance(val_str, str):
        val_str = val_str.replace(',', '.')
    v = float(val_str)
    if not math.isfinite(v):
        raise ValueError(f"Value '{val_str}' is not a finite float.")
    return v

# Classe responsável por rodar a simulação em uma thread separada para não travar a interface
class SimulationThread(QThread):
    finished = pyqtSignal(object)

    # Construtor da thread de simulação
    def __init__(self, state):
        super().__init__()
        self.state = state

    # Função principal que executa ao iniciar a thread
    def run(self):
        engine = PowerSystemEngine()
        engine.build_network(self.state)
        engine.run_power_flow()

        self.finished.emit(engine.results)

# Classe responsável por rodar a simulação de curva PV em uma thread separada
class PVCurveThread(QThread):
    finished = pyqtSignal(object)

    # Construtor da thread de Curva PV
    def __init__(self, state, target_bus):
        super().__init__()
        self.state = state
        self.target_bus = target_bus

    # Função principal que executa a simulação PV ao iniciar a thread
    def run(self):
        engine = PowerSystemEngine()
        engine.build_network(self.state)
        engine.run_power_flow()

        if self.target_bus:
            engine.generate_pv_curve(self.target_bus)
        self.finished.emit(engine.results)

# Classe principal que integra a UI com a Engine e gerencia o estado da aplicação
class MainController:
    # Construtor do Controlador Principal
    def __init__(self):
        # Definir AppUserModelID para ícone da barra de tarefas do Windows
        if sys.platform == 'win32':
            myappid = 'tcc.ll.13bus.1.0' # string arbitrária
            ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID(myappid)
            
        self.app = QApplication(sys.argv)
        self.app.setWindowIcon(QIcon(resource_path("Logo.png")))
        self.ui = MainWindowUI()
        self.ui.setWindowIcon(QIcon(resource_path("Logo.png")))
        self.state = SystemState()
        self.scenarios = {}

        self.init_default_data()
        self.load_settings()
        self.load_scenarios()
        self.setup_connections()

        # Desenho inicial
        self.update_diagram()
        self.populate_params_tables()
        self.populate_target_combo()
        self.update_scenarios_combo()


    # Salva as configurações de parâmetros no QSettings do sistema
    def save_settings(self):
        settings = QSettings("SimuladorSEP", "Parametros")

        # Salvar barras
        buses_data = {}
        for bus_id, bus in self.state.buses.items():
            buses_data[bus_id] = {
                "p_load_kw": bus.p_load_kw,
                "q_load_kvar": bus.q_load_kvar,
                "p_gen_kw": bus.p_gen_kw
            }
        settings.setValue("buses", json.dumps(buses_data))

        # Salvar linhas
        lines_data = {}
        for line_id, line in self.state.lines.items():
            lines_data[line_id] = {
                "r_ohm_per_km": line.r_ohm_per_km,
                "x_ohm_per_km": line.x_ohm_per_km,
                "length_km": line.length_km,
                "sn_mva": line.sn_mva,
                "vk_percent": line.vk_percent,
                "vkr_percent": line.vkr_percent,
                "pfe_kw": line.pfe_kw,
                "i0_percent": line.i0_percent
            }
        settings.setValue("lines", json.dumps(lines_data))

        # Salvar cabos
        cables_data = {}
        for c_name, cable in self.state.cables.items():
            cables_data[c_name] = {
                "r_ohm_per_km": cable.r_ohm_per_km,
                "x_ohm_per_km": cable.x_ohm_per_km
            }
        settings.setValue("cables", json.dumps(cables_data))

    # Carrega as configurações de parâmetros salvas anteriormente
    def load_settings(self):
        settings = QSettings("SimuladorSEP", "Parametros")

        buses_str = settings.value("buses", "")
        if buses_str:
            try:
                buses_data = json.loads(buses_str)
                for bus_id_str, data in buses_data.items():
                    bus_id = int(bus_id_str)
                    if bus_id in self.state.buses:
                        self.state.buses[bus_id].p_load_kw = float(data.get("p_load_kw", self.state.buses[bus_id].p_load_kw))
                        self.state.buses[bus_id].q_load_kvar = float(data.get("q_load_kvar", self.state.buses[bus_id].q_load_kvar))
                        self.state.buses[bus_id].p_gen_kw = float(data.get("p_gen_kw", self.state.buses[bus_id].p_gen_kw))
            except Exception as e:
                print("Erro ao carregar parâmetros de barras:", e)

        lines_str = settings.value("lines", "")
        if lines_str:
            try:
                lines_data = json.loads(lines_str)
                for line_id_str, data in lines_data.items():
                    line_id = int(line_id_str)
                    if line_id in self.state.lines:
                        self.state.lines[line_id].r_ohm_per_km = float(data.get("r_ohm_per_km", self.state.lines[line_id].r_ohm_per_km))
                        self.state.lines[line_id].x_ohm_per_km = float(data.get("x_ohm_per_km", self.state.lines[line_id].x_ohm_per_km))
                        self.state.lines[line_id].length_km = float(data.get("length_km", self.state.lines[line_id].length_km))
                        self.state.lines[line_id].sn_mva = float(data.get("sn_mva", self.state.lines[line_id].sn_mva))
                        self.state.lines[line_id].vk_percent = float(data.get("vk_percent", self.state.lines[line_id].vk_percent))
                        self.state.lines[line_id].vkr_percent = float(data.get("vkr_percent", self.state.lines[line_id].vkr_percent))
                        self.state.lines[line_id].pfe_kw = float(data.get("pfe_kw", self.state.lines[line_id].pfe_kw))
                        self.state.lines[line_id].i0_percent = float(data.get("i0_percent", self.state.lines[line_id].i0_percent))
            except Exception as e:
                print("Erro ao carregar parâmetros de linhas:", e)
                
        cables_str = settings.value("cables", "")
        if cables_str:
            try:
                cables_data = json.loads(cables_str)
                for c_name, data in cables_data.items():
                    self.state.cables[c_name] = CableConfig(
                        name=c_name,
                        r_ohm_per_km=float(data.get("r_ohm_per_km", 0.1)),
                        x_ohm_per_km=float(data.get("x_ohm_per_km", 0.1))
                    )
            except Exception as e:
                print("Erro ao carregar parâmetros de cabos:", e)

    # Carrega os cenários predefinidos e os cenários do usuário
    def load_scenarios(self):
        default_scenarios = {
            "TCC Ref": {"buses": {"650": {"p_load_kw": 0.0, "q_load_kvar": 0.0, "p_gen_kw": 0.0}, "632": {"p_load_kw": 0.0, "q_load_kvar": 0.0, "p_gen_kw": 0.0}, "645": {"p_load_kw": 170.0, "q_load_kvar": 125.0, "p_gen_kw": 0.0}, "646": {"p_load_kw": 230.0, "q_load_kvar": 132.0, "p_gen_kw": 0.0}, "633": {"p_load_kw": 0.0, "q_load_kvar": 0.0, "p_gen_kw": 0.0}, "634": {"p_load_kw": 340.0, "q_load_kvar": 120.0, "p_gen_kw": 0.0}, "671": {"p_load_kw": 1155.0, "q_load_kvar": 660.0, "p_gen_kw": 0.0}, "684": {"p_load_kw": 0.0, "q_load_kvar": 0.0, "p_gen_kw": 0.0}, "611": {"p_load_kw": 170.0, "q_load_kvar": -220.0, "p_gen_kw": 0.0}, "652": {"p_load_kw": 128.0, "q_load_kvar": 86.0, "p_gen_kw": 0.0}, "692": {"p_load_kw": 170.0, "q_load_kvar": 151.0, "p_gen_kw": 0.0}, "675": {"p_load_kw": 843.0, "q_load_kvar": -138.0, "p_gen_kw": 0.0}, "680": {"p_load_kw": 0.0, "q_load_kvar": 0.0, "p_gen_kw": 0.0}}, "lines": {"1": {"r_ohm_per_km": 0.1155, "x_ohm_per_km": 0.371, "length_km": 0.6096, "sn_mva": 5.0, "vk_percent": 5.0, "vkr_percent": 1.0, "pfe_kw": 10.0, "i0_percent": 0.5}, "2": {"r_ohm_per_km": 0.3679, "x_ohm_per_km": 0.4726, "length_km": 0.1524, "sn_mva": 5.0, "vk_percent": 5.0, "vkr_percent": 1.0, "pfe_kw": 10.0, "i0_percent": 0.5}, "3": {"r_ohm_per_km": 0.3679, "x_ohm_per_km": 0.4726, "length_km": 0.0914, "sn_mva": 5.0, "vk_percent": 5.0, "vkr_percent": 1.0, "pfe_kw": 10.0, "i0_percent": 0.5}, "4": {"r_ohm_per_km": 0.3679, "x_ohm_per_km": 0.4726, "length_km": 0.1524, "sn_mva": 5.0, "vk_percent": 5.0, "vkr_percent": 1.0, "pfe_kw": 10.0, "i0_percent": 0.5}, "5": {"r_ohm_per_km": 0.001, "x_ohm_per_km": 0.001, "length_km": 0.001, "sn_mva": 0.5, "vk_percent": 4.0, "vkr_percent": 1.0, "pfe_kw": 0.0, "i0_percent": 0.5}, "6": {"r_ohm_per_km": 0.1155, "x_ohm_per_km": 0.371, "length_km": 0.6096, "sn_mva": 5.0, "vk_percent": 5.0, "vkr_percent": 1.0, "pfe_kw": 10.0, "i0_percent": 0.5}, "7": {"r_ohm_per_km": 0.3679, "x_ohm_per_km": 0.4726, "length_km": 0.0914, "sn_mva": 5.0, "vk_percent": 5.0, "vkr_percent": 1.0, "pfe_kw": 10.0, "i0_percent": 0.5}, "8": {"r_ohm_per_km": 0.3679, "x_ohm_per_km": 0.4726, "length_km": 0.0914, "sn_mva": 5.0, "vk_percent": 5.0, "vkr_percent": 1.0, "pfe_kw": 10.0, "i0_percent": 0.5}, "9": {"r_ohm_per_km": 0.3679, "x_ohm_per_km": 0.4726, "length_km": 0.2438, "sn_mva": 5.0, "vk_percent": 5.0, "vkr_percent": 1.0, "pfe_kw": 10.0, "i0_percent": 0.5}, "10": {"r_ohm_per_km": 0.001, "x_ohm_per_km": 0.001, "length_km": 0.001, "sn_mva": 5.0, "vk_percent": 5.0, "vkr_percent": 1.0, "pfe_kw": 10.0, "i0_percent": 0.5}, "11": {"r_ohm_per_km": 0.3679, "x_ohm_per_km": 0.4726, "length_km": 0.1524, "sn_mva": 5.0, "vk_percent": 5.0, "vkr_percent": 1.0, "pfe_kw": 10.0, "i0_percent": 0.5}, "12": {"r_ohm_per_km": 0.1155, "x_ohm_per_km": 0.371, "length_km": 0.3048, "sn_mva": 5.0, "vk_percent": 5.0, "vkr_percent": 1.0, "pfe_kw": 10.0, "i0_percent": 0.5}}},
            "Cenário Base": {"buses": {"650": {"p_load_kw": 0.0, "q_load_kvar": 0.0, "p_gen_kw": 0.0}, "632": {"p_load_kw": 0.0, "q_load_kvar": 0.0, "p_gen_kw": 0.0}, "645": {"p_load_kw": 200.0, "q_load_kvar": 95.0, "p_gen_kw": 0.0}, "646": {"p_load_kw": 250.0, "q_load_kvar": 120.0, "p_gen_kw": 0.0}, "633": {"p_load_kw": 0.0, "q_load_kvar": 0.0, "p_gen_kw": 0.0}, "634": {"p_load_kw": 300.0, "q_load_kvar": 120.0, "p_gen_kw": 0.0}, "671": {"p_load_kw": 800.0, "q_load_kvar": 380.0, "p_gen_kw": 0.0}, "684": {"p_load_kw": 350.0, "q_load_kvar": 170.0, "p_gen_kw": 0.0}, "611": {"p_load_kw": 250.0, "q_load_kvar": 120.0, "p_gen_kw": 0.0}, "652": {"p_load_kw": 150.0, "q_load_kvar": 70.0, "p_gen_kw": 0.0}, "692": {"p_load_kw": 170.0, "q_load_kvar": 80.0, "p_gen_kw": 0.0}, "675": {"p_load_kw": 450.0, "q_load_kvar": 210.0, "p_gen_kw": 0.0}, "680": {"p_load_kw": 500.0, "q_load_kvar": 240.0, "p_gen_kw": 0.0}}, "lines": {"1": {"r_ohm_per_km": 0.17, "x_ohm_per_km": 0.36, "length_km": 0.6096, "sn_mva": 5.0, "vk_percent": 5.0, "vkr_percent": 1.0, "pfe_kw": 10.0, "i0_percent": 0.5}, "2": {"r_ohm_per_km": 0.88, "x_ohm_per_km": 0.42, "length_km": 0.1524, "sn_mva": 5.0, "vk_percent": 5.0, "vkr_percent": 1.0, "pfe_kw": 10.0, "i0_percent": 0.5}, "3": {"r_ohm_per_km": 0.88, "x_ohm_per_km": 0.42, "length_km": 0.0914, "sn_mva": 5.0, "vk_percent": 5.0, "vkr_percent": 1.0, "pfe_kw": 10.0, "i0_percent": 0.5}, "4": {"r_ohm_per_km": 0.27, "x_ohm_per_km": 0.38, "length_km": 0.1524, "sn_mva": 5.0, "vk_percent": 5.0, "vkr_percent": 1.0, "pfe_kw": 10.0, "i0_percent": 0.5}, "5": {"r_ohm_per_km": 0.001, "x_ohm_per_km": 0.001, "length_km": 0.001, "sn_mva": 0.5, "vk_percent": 4.0, "vkr_percent": 1.0, "pfe_kw": 0.0, "i0_percent": 0.5}, "6": {"r_ohm_per_km": 0.17, "x_ohm_per_km": 0.36, "length_km": 0.6096, "sn_mva": 5.0, "vk_percent": 5.0, "vkr_percent": 1.0, "pfe_kw": 10.0, "i0_percent": 0.5}, "7": {"r_ohm_per_km": 0.27, "x_ohm_per_km": 0.38, "length_km": 0.0914, "sn_mva": 5.0, "vk_percent": 5.0, "vkr_percent": 1.0, "pfe_kw": 10.0, "i0_percent": 0.5}, "8": {"r_ohm_per_km": 0.88, "x_ohm_per_km": 0.42, "length_km": 0.0914, "sn_mva": 5.0, "vk_percent": 5.0, "vkr_percent": 1.0, "pfe_kw": 10.0, "i0_percent": 0.5}, "9": {"r_ohm_per_km": 0.88, "x_ohm_per_km": 0.42, "length_km": 0.2438, "sn_mva": 5.0, "vk_percent": 5.0, "vkr_percent": 1.0, "pfe_kw": 10.0, "i0_percent": 0.5}, "10": {"r_ohm_per_km": 0.001, "x_ohm_per_km": 0.001, "length_km": 0.001, "sn_mva": 5.0, "vk_percent": 5.0, "vkr_percent": 1.0, "pfe_kw": 10.0, "i0_percent": 0.5}, "11": {"r_ohm_per_km": 0.27, "x_ohm_per_km": 0.38, "length_km": 0.1524, "sn_mva": 5.0, "vk_percent": 5.0, "vkr_percent": 1.0, "pfe_kw": 10.0, "i0_percent": 0.5}, "12": {"r_ohm_per_km": 0.17, "x_ohm_per_km": 0.36, "length_km": 0.3048, "sn_mva": 5.0, "vk_percent": 5.0, "vkr_percent": 1.0, "pfe_kw": 10.0, "i0_percent": 0.5}}}
        }

        settings = QSettings("SimuladorSEP", "Cenarios")
        data = settings.value("data", None)
        
        if data is None or data == "{}":
            self.scenarios = default_scenarios
        else:
            try:
                self.scenarios = json.loads(data)
            except:
                self.scenarios = default_scenarios

    # Salva todos os cenários no QSettings do sistema
    def save_scenarios(self):
        settings = QSettings("SimuladorSEP", "Cenarios")
        settings.setValue("data", json.dumps(self.scenarios))
        self.update_scenarios_combo()

    # Atualiza a lista da combobox com os cenários disponíveis
    def update_scenarios_combo(self):
        self.ui.combo_scenarios.blockSignals(True)
        self.ui.combo_scenarios.clear()
        self.ui.combo_scenarios.addItem("-- Selecionar Cenário --")
        self.ui.combo_scenarios.addItems(list(self.scenarios.keys()))
        self.ui.combo_scenarios.blockSignals(False)

    # Evento acionado quando o usuário seleciona um cenário na combobox
    def on_scenario_selected(self, idx):
        if idx <= 0:
            return
        name = self.ui.combo_scenarios.currentText()
        if name in self.scenarios:
            scen_data = self.scenarios[name]
            self.apply_scenario_data(scen_data)
            self.populate_params_tables()
            self.update_diagram()
            self.save_settings()
            self.ui.toast.show_toast(f"Cenário '{name}' carregado!", True, self.ui)

    # Aplica os dados de um cenário específico ao estado atual do sistema
    def apply_scenario_data(self, scen_data):
        buses_data = scen_data.get("buses", {})
        for bus_id_str, data in buses_data.items():
            bus_id = int(bus_id_str)
            if bus_id in self.state.buses:
                self.state.buses[bus_id].p_load_kw = float(data.get("p_load_kw", self.state.buses[bus_id].p_load_kw))
                self.state.buses[bus_id].q_load_kvar = float(data.get("q_load_kvar", self.state.buses[bus_id].q_load_kvar))
                self.state.buses[bus_id].p_gen_kw = float(data.get("p_gen_kw", self.state.buses[bus_id].p_gen_kw))

        lines_data = scen_data.get("lines", {})
        for line_id_str, data in lines_data.items():
            line_id = int(line_id_str)
            if line_id in self.state.lines:
                self.state.lines[line_id].r_ohm_per_km = float(data.get("r_ohm_per_km", self.state.lines[line_id].r_ohm_per_km))
                self.state.lines[line_id].x_ohm_per_km = float(data.get("x_ohm_per_km", self.state.lines[line_id].x_ohm_per_km))
                self.state.lines[line_id].length_km = float(data.get("length_km", self.state.lines[line_id].length_km))
                self.state.lines[line_id].sn_mva = float(data.get("sn_mva", self.state.lines[line_id].sn_mva))
                self.state.lines[line_id].vk_percent = float(data.get("vk_percent", self.state.lines[line_id].vk_percent))
                self.state.lines[line_id].vkr_percent = float(data.get("vkr_percent", self.state.lines[line_id].vkr_percent))
                self.state.lines[line_id].pfe_kw = float(data.get("pfe_kw", self.state.lines[line_id].pfe_kw))
                self.state.lines[line_id].i0_percent = float(data.get("i0_percent", self.state.lines[line_id].i0_percent))

    # Ação disparada pelo botão de salvar um novo cenário
    def save_scenario_action(self):
        self.save_params()  # Salva alterações antes de salvar o cenário
        from PyQt6.QtWidgets import QInputDialog
        name, ok = QInputDialog.getText(self.ui, "Salvar Cenário", "Nome do Cenário:")
        if ok and name.strip():
            name = name.strip()
            # Serializar parâmetros atuais
            buses_data = {}
            for bus_id, bus in self.state.buses.items():
                buses_data[bus_id] = {
                    "p_load_kw": bus.p_load_kw,
                    "q_load_kvar": bus.q_load_kvar,
                    "p_gen_kw": bus.p_gen_kw
                }
            lines_data = {}
            for line_id, line in self.state.lines.items():
                lines_data[line_id] = {
                    "r_ohm_per_km": line.r_ohm_per_km,
                    "x_ohm_per_km": line.x_ohm_per_km,
                    "length_km": line.length_km,
                    "sn_mva": line.sn_mva,
                    "vk_percent": line.vk_percent,
                    "vkr_percent": line.vkr_percent,
                    "pfe_kw": line.pfe_kw,
                    "i0_percent": line.i0_percent
                }
            self.scenarios[name] = {"buses": buses_data, "lines": lines_data}
            self.save_scenarios()
            self.ui.toast.show_toast(f"Cenário '{name}' salvo!", True, self.ui)

    # Ação disparada pelo botão de excluir um cenário existente
    def delete_scenario_action(self):
        from PyQt6.QtWidgets import QInputDialog, QMessageBox
        if not self.scenarios:
            QMessageBox.information(self.ui, "Excluir Cenário", "Nenhum cenário salvo.")
            return

        items = list(self.scenarios.keys())
        item, ok = QInputDialog.getItem(self.ui, "Excluir Cenário", "Selecione o cenário a excluir:", items, 0, False)
        if ok and item:
            del self.scenarios[item]
            self.save_scenarios()
            self.ui.toast.show_toast(f"Cenário '{item}' excluído!", True, self.ui)

    # Ação disparada pelo botão de renomear um cenário existente
    def rename_scenario_action(self):
        from PyQt6.QtWidgets import QInputDialog, QMessageBox
        if not self.scenarios:
            QMessageBox.information(self.ui, "Renomear Cenário", "Nenhum cenário salvo.")
            return

        items = list(self.scenarios.keys())
        item, ok = QInputDialog.getItem(self.ui, "Renomear Cenário", "Selecione o cenário a renomear:", items, 0, False)
        if ok and item:
            new_name, ok2 = QInputDialog.getText(self.ui, "Renomear Cenário", "Novo nome para o cenário:", text=item)
            if ok2 and new_name.strip() and new_name.strip() != item:
                new_name = new_name.strip()
                if new_name in self.scenarios:
                    QMessageBox.warning(self.ui, "Erro", "Já existe um cenário com esse nome.")
                    return
                # Substituir a chave
                scenario_data = self.scenarios.pop(item)
                self.scenarios[new_name] = scenario_data
                self.save_scenarios()
                self.ui.toast.show_toast(f"Cenário renomeado para '{new_name}'!", True, self.ui)

    # Inicializa os dados padrões do sistema IEEE de 13 barras
    def init_default_data(self):
        # Sistema de 13 barras baseado no layout do IEEE 13 Node Test Feeder
        self.state.buses[650] = BusData(id=650, name="650 (Slack)", vn_kv=13.8, type='slack', v_target_pu=1.0)
        self.state.buses[632] = BusData(id=632, name="632", vn_kv=13.8, type='pq', p_load_kw=0, q_load_kvar=0)
        self.state.buses[645] = BusData(id=645, name="645", vn_kv=13.8, type='pq', p_load_kw=170, q_load_kvar=125)
        self.state.buses[646] = BusData(id=646, name="646", vn_kv=13.8, type='pq', p_load_kw=230, q_load_kvar=132)
        self.state.buses[633] = BusData(id=633, name="633", vn_kv=13.8, type='pq', p_load_kw=0, q_load_kvar=0)
        self.state.buses[634] = BusData(id=634, name="634", vn_kv=0.22, type='pq', p_load_kw=340, q_load_kvar=120)
        self.state.buses[671] = BusData(id=671, name="671", vn_kv=13.8, type='pq', p_load_kw=1155, q_load_kvar=660)
        self.state.buses[684] = BusData(id=684, name="684", vn_kv=13.8, type='pq', p_load_kw=0, q_load_kvar=0)
        self.state.buses[611] = BusData(id=611, name="611", vn_kv=13.8, type='pq', p_load_kw=170, q_load_kvar=-220)
        self.state.buses[652] = BusData(id=652, name="652", vn_kv=13.8, type='pq', p_load_kw=128, q_load_kvar=86)
        self.state.buses[692] = BusData(id=692, name="692", vn_kv=13.8, type='pq', p_load_kw=170, q_load_kvar=151)
        self.state.buses[675] = BusData(id=675, name="675", vn_kv=13.8, type='pq', p_load_kw=843, q_load_kvar=-138)
        self.state.buses[680] = BusData(id=680, name="680", vn_kv=13.8, type='pq', p_load_kw=0, q_load_kvar=0)

        # Conectar as barras
        # (id, from, to, length, r, x)
        self.state.lines[1] = LineData(1, 650, 632, 0.6096, 0.1155, 0.371, is_transformer=False)
        self.state.lines[2] = LineData(2, 632, 645, 0.1524, 0.3679, 0.4726)
        self.state.lines[3] = LineData(3, 645, 646, 0.0914, 0.3679, 0.4726)
        self.state.lines[4] = LineData(4, 632, 633, 0.1524, 0.3679, 0.4726)
        self.state.lines[5] = LineData(5, 633, 634, 0.0, 0.0, 0.0, is_transformer=True, sn_mva=0.5, vk_percent=4.0, vkr_percent=1.0, pfe_kw=0.0, i0_percent=0.5)
        self.state.lines[6] = LineData(6, 632, 671, 0.6096, 0.1155, 0.371)
        self.state.lines[7] = LineData(7, 671, 684, 0.0914, 0.3679, 0.4726)
        self.state.lines[8] = LineData(8, 684, 611, 0.0914, 0.3679, 0.4726)
        self.state.lines[9] = LineData(9, 684, 652, 0.2438, 0.3679, 0.4726)
        self.state.lines[10] = LineData(10, 671, 692, 0.01, 0.01, 0.01) # chave (evitar divisão por zero e mau condicionamento)
        self.state.lines[11] = LineData(11, 692, 675, 0.1524, 0.3679, 0.4726)
        self.state.lines[12] = LineData(12, 671, 680, 0.3048, 0.1155, 0.371)

        self.populate_params_tables()

    # Estabelece as conexões de sinais (signals) e slots da UI
    def setup_connections(self):
        self.ui.btn_simulate.clicked.connect(self.run_simulation)
        self.ui.btn_export_image.clicked.connect(self.export_image)
        self.ui.btn_simulate_pv.clicked.connect(self.run_pv_simulation)
        self.ui.diagram_view.data_updated.connect(self.on_diagram_data_updated)
        self.ui.btn_save_params.clicked.connect(self.save_params)
        self.ui.btn_export_params.clicked.connect(self.export_params)
        self.ui.btn_import_params.clicked.connect(self.import_params)
        self.ui.btn_export_results.clicked.connect(self.export_results)
        self.ui.btn_export_plot.clicked.connect(self.export_plot)
        self.ui.app_closed.connect(self.save_settings)
        self.ui.btn_save_scenario.clicked.connect(self.save_scenario_action)
        self.ui.btn_rename_scenario.clicked.connect(self.rename_scenario_action)
        self.ui.btn_delete_scenario.clicked.connect(self.delete_scenario_action)
        self.ui.combo_scenarios.currentIndexChanged.connect(self.on_scenario_selected)
        
        self.ui.btn_add_cable.clicked.connect(self.add_cable)
        self.ui.btn_remove_cable.clicked.connect(self.remove_cable)

    # Callback invocado quando um nó do diagrama é alterado
    def on_diagram_data_updated(self):
        self.update_diagram()
        self.populate_params_tables()
        self.save_settings()
        
    # Ação para adicionar um novo cabo personalizado
    def add_cable(self):
        from PyQt6.QtWidgets import QInputDialog
        name, ok = QInputDialog.getText(self.ui, "Novo Cabo", "Nome do Cabo:")
        if ok and name and name not in self.state.cables:
            self.state.cables[name] = CableConfig(name=name, r_ohm_per_km=0.1, x_ohm_per_km=0.1)
            self.populate_params_tables()
            self.save_settings()
            
    # Ação para remover um cabo selecionado
    def remove_cable(self):
        from PyQt6.QtWidgets import QInputDialog
        if not self.state.cables:
            return
        items = list(self.state.cables.keys())
        item, ok = QInputDialog.getItem(self.ui, "Remover Cabo", "Selecione o cabo:", items, 0, False)
        if ok and item in self.state.cables:
            del self.state.cables[item]
            self.populate_params_tables()
            self.save_settings()

    # Preenche todas as tabelas de parâmetros com os dados do SystemState
    def populate_params_tables(self):
        from PyQt6.QtWidgets import QTableWidgetItem
        # Preencher Cabos
        self.ui.table_cables.setRowCount(len(self.state.cables))
        self.ui.table_cables.setColumnCount(3)
        self.ui.table_cables.setHorizontalHeaderLabels(["Nome", "R (ohm/km)", "X (ohm/km)"])
        for i, (name, cable) in enumerate(self.state.cables.items()):
            self.ui.table_cables.setItem(i, 0, QTableWidgetItem(name))
            self.ui.table_cables.setItem(i, 1, QTableWidgetItem(str(cable.r_ohm_per_km)))
            self.ui.table_cables.setItem(i, 2, QTableWidgetItem(str(cable.x_ohm_per_km)))

        # Preencher Barras
        self.ui.table_params_buses.setRowCount(len(self.state.buses))
        self.ui.table_params_buses.setColumnCount(4)
        self.ui.table_params_buses.setHorizontalHeaderLabels(["ID", "P Load (kW)", "Q Load (kVAr)", "Geração (kW)"])
        for i, (bus_id, bus) in enumerate(self.state.buses.items()):
            self.ui.table_params_buses.setItem(i, 0, QTableWidgetItem(str(bus.id)))
            self.ui.table_params_buses.setItem(i, 1, QTableWidgetItem(str(bus.p_load_kw)))
            self.ui.table_params_buses.setItem(i, 2, QTableWidgetItem(str(bus.q_load_kvar)))
            self.ui.table_params_buses.setItem(i, 3, QTableWidgetItem(str(bus.p_gen_kw)))

        # Preencher Linhas
        from PyQt6.QtWidgets import QComboBox
        normal_lines = [l for l in self.state.lines.values() if not l.is_transformer]
        self.ui.table_params_lines.setRowCount(len(normal_lines))
        self.ui.table_params_lines.setColumnCount(5)
        self.ui.table_params_lines.setHorizontalHeaderLabels(["ID", "R (ohm/km)", "X (ohm/km)", "Length (km)", "Cabo"])
        for i, line in enumerate(normal_lines):
            item = QTableWidgetItem(f"{line.from_bus} - {line.to_bus}")
            item.setData(Qt.ItemDataRole.UserRole, line.id)
            self.ui.table_params_lines.setItem(i, 0, item)
            self.ui.table_params_lines.setItem(i, 1, QTableWidgetItem(str(line.r_ohm_per_km)))
            self.ui.table_params_lines.setItem(i, 2, QTableWidgetItem(str(line.x_ohm_per_km)))
            self.ui.table_params_lines.setItem(i, 3, QTableWidgetItem(str(line.length_km)))
            
            combo = QComboBox()
            combo.addItem("Personalizado")
            for c_name in self.state.cables.keys():
                combo.addItem(c_name)
            
            # Selecionar o cabo correto se houver correspondência
            for c_name, c_data in self.state.cables.items():
                if abs(c_data.r_ohm_per_km - line.r_ohm_per_km) < 1e-4 and abs(c_data.x_ohm_per_km - line.x_ohm_per_km) < 1e-4:
                    combo.setCurrentText(c_name)
                    break
            
            # Conectar
            def on_cable_selected(text, row=i):
                if text in self.state.cables:
                    c = self.state.cables[text]
                    self.ui.table_params_lines.item(row, 1).setText(str(c.r_ohm_per_km))
                    self.ui.table_params_lines.item(row, 2).setText(str(c.x_ohm_per_km))
            
            combo.currentTextChanged.connect(on_cable_selected)
            self.ui.table_params_lines.setCellWidget(i, 4, combo)

        # Preencher Transformadores
        trafos = [l for l in self.state.lines.values() if l.is_transformer]
        self.ui.table_params_trafos.setRowCount(len(trafos))
        self.ui.table_params_trafos.setColumnCount(6)
        self.ui.table_params_trafos.setHorizontalHeaderLabels(["ID", "Sn (MVA)", "vk (%)", "vkr (%)", "pfe (kW)", "i0 (%)"])
        for i, trafo in enumerate(trafos):
            item = QTableWidgetItem(f"{trafo.from_bus} - {trafo.to_bus}")
            item.setData(Qt.ItemDataRole.UserRole, trafo.id)
            self.ui.table_params_trafos.setItem(i, 0, item)
            self.ui.table_params_trafos.setItem(i, 1, QTableWidgetItem(str(trafo.sn_mva)))
            self.ui.table_params_trafos.setItem(i, 2, QTableWidgetItem(str(trafo.vk_percent)))
            self.ui.table_params_trafos.setItem(i, 3, QTableWidgetItem(str(trafo.vkr_percent)))
            self.ui.table_params_trafos.setItem(i, 4, QTableWidgetItem(str(trafo.pfe_kw)))
            self.ui.table_params_trafos.setItem(i, 5, QTableWidgetItem(str(trafo.i0_percent)))
            
        self.ui.table_cables.resizeColumnsToContents()
        self.ui.table_params_buses.resizeColumnsToContents()
        self.ui.table_params_lines.resizeColumnsToContents()
        self.ui.table_params_trafos.resizeColumnsToContents()
        
        self.adjust_table_size(self.ui.table_cables)
        self.adjust_table_size(self.ui.table_params_buses)
        self.adjust_table_size(self.ui.table_params_lines)
        self.adjust_table_size(self.ui.table_params_trafos)

    # Ajusta as alturas e larguras mínimas das tabelas da interface
    def adjust_table_size(self, table):
        h = table.horizontalHeader().height()
        for i in range(table.rowCount()):
            h += table.rowHeight(i)
        table.setMinimumHeight(h + 10)
        
        w = table.verticalHeader().width() if table.verticalHeader().isVisible() else 0
        for i in range(table.columnCount()):
            w += table.columnWidth(i)
        table.setMinimumWidth(w + 30)

    # Salva as modificações feitas diretamente nas tabelas de parâmetros para o SystemState
    def save_params(self):
        try:
            for i in range(self.ui.table_cables.rowCount()):
                name = self.ui.table_cables.item(i, 0).text()
                if name in self.state.cables:
                    self.state.cables[name].r_ohm_per_km = safe_float(self.ui.table_cables.item(i, 1).text())
                    self.state.cables[name].x_ohm_per_km = safe_float(self.ui.table_cables.item(i, 2).text())
                    
            for i in range(self.ui.table_params_buses.rowCount()):
                bus_id = int(self.ui.table_params_buses.item(i, 0).text())
                if bus_id in self.state.buses:
                    self.state.buses[bus_id].p_load_kw = safe_float(self.ui.table_params_buses.item(i, 1).text())
                    self.state.buses[bus_id].q_load_kvar = safe_float(self.ui.table_params_buses.item(i, 2).text())
                    self.state.buses[bus_id].p_gen_kw = safe_float(self.ui.table_params_buses.item(i, 3).text())

            for i in range(self.ui.table_params_lines.rowCount()):
                line_id = self.ui.table_params_lines.item(i, 0).data(Qt.ItemDataRole.UserRole)
                if line_id in self.state.lines:
                    self.state.lines[line_id].r_ohm_per_km = safe_float(self.ui.table_params_lines.item(i, 1).text())
                    self.state.lines[line_id].x_ohm_per_km = safe_float(self.ui.table_params_lines.item(i, 2).text())
                    self.state.lines[line_id].length_km = safe_float(self.ui.table_params_lines.item(i, 3).text())

            for i in range(self.ui.table_params_trafos.rowCount()):
                trafo_id = self.ui.table_params_trafos.item(i, 0).data(Qt.ItemDataRole.UserRole)
                if trafo_id in self.state.lines:
                    self.state.lines[trafo_id].sn_mva = safe_float(self.ui.table_params_trafos.item(i, 1).text())
                    self.state.lines[trafo_id].vk_percent = safe_float(self.ui.table_params_trafos.item(i, 2).text())
                    self.state.lines[trafo_id].vkr_percent = safe_float(self.ui.table_params_trafos.item(i, 3).text())
                    self.state.lines[trafo_id].pfe_kw = safe_float(self.ui.table_params_trafos.item(i, 4).text())
                    self.state.lines[trafo_id].i0_percent = safe_float(self.ui.table_params_trafos.item(i, 5).text())

            self.update_diagram()
            self.save_settings()
            QMessageBox.information(self.ui, "Sucesso", "Parâmetros salvos com sucesso!")
        except ValueError:
            QMessageBox.warning(self.ui, "Erro", "Valores inválidos inseridos. Use apenas números.")

    # Exporta os parâmetros atuais para um arquivo Excel
    def export_params(self):
        import datetime
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        default_name = f"parametros_{timestamp}.xlsx"
        filename, _ = QFileDialog.getSaveFileName(self.ui, "Exportar Parâmetros", default_name, "Excel Files (*.xlsx)")
        if not filename:
            return

        try:
            wb = Workbook()
            ws = wb.active

            ws.append(["[Buses]"])
            ws.append(["ID", "P Load (kW)", "Q Load (kVAr)", "Geração (kW)"])
            for bus_id, bus in self.state.buses.items():
                ws.append([bus_id, bus.p_load_kw, bus.q_load_kvar, bus.p_gen_kw])

            ws.append([])
            ws.append(["[Lines]"])
            ws.append(["ID", "R (ohm/km)", "X (ohm/km)", "Length (km)"])
            for line_id, line in self.state.lines.items():
                if not line.is_transformer:
                    ws.append([line.id, line.r_ohm_per_km, line.x_ohm_per_km, line.length_km])

            ws.append([])
            ws.append(["[Transformers]"])
            ws.append(["ID", "Sn (MVA)", "vk (%)", "vkr (%)", "pfe (kW)", "i0 (%)"])
            for line_id, line in self.state.lines.items():
                if line.is_transformer:
                    ws.append([line.id, line.sn_mva, line.vk_percent, line.vkr_percent, line.pfe_kw, line.i0_percent])

            wb.save(filename)
            QMessageBox.information(self.ui, "Sucesso", "Parâmetros exportados com sucesso!")
        except Exception as e:
            QMessageBox.critical(self.ui, "Erro", f"Erro ao exportar: {str(e)}")

    # Importa parâmetros para o sistema a partir de um arquivo Excel
    def import_params(self):
        filename, _ = QFileDialog.getOpenFileName(self.ui, "Importar Parâmetros", "", "Excel Files (*.xlsx)")
        if not filename:
            return

        try:
            wb = load_workbook(filename)
            ws = wb.active
            mode = None

            for row in ws.iter_rows(values_only=True):
                if not row or row[0] is None:
                    continue
                if row[0] == "[Buses]":
                    mode = "buses"
                    continue
                elif row[0] == "[Lines]":
                    mode = "lines"
                    continue
                elif row[0] == "[Transformers]":
                    mode = "transformers"
                    continue

                if row[0] == "ID":
                    continue

                if mode == "buses":
                    bus_id = int(row[0])
                    if bus_id in self.state.buses:
                        self.state.buses[bus_id].p_load_kw = safe_float(row[1])
                        self.state.buses[bus_id].q_load_kvar = safe_float(row[2])
                        self.state.buses[bus_id].p_gen_kw = safe_float(row[3])
                elif mode == "lines":
                    line_id = int(row[0])
                    if line_id in self.state.lines:
                        self.state.lines[line_id].r_ohm_per_km = safe_float(row[1])
                        self.state.lines[line_id].x_ohm_per_km = safe_float(row[2])
                        self.state.lines[line_id].length_km = safe_float(row[3])
                elif mode == "transformers":
                    trafo_id = int(row[0])
                    if trafo_id in self.state.lines:
                        self.state.lines[trafo_id].sn_mva = safe_float(row[1])
                        self.state.lines[trafo_id].vk_percent = safe_float(row[2])
                        self.state.lines[trafo_id].vkr_percent = safe_float(row[3])
                        self.state.lines[trafo_id].pfe_kw = safe_float(row[4])
                        self.state.lines[trafo_id].i0_percent = safe_float(row[5])

            self.populate_params_tables()
            self.update_diagram()
            self.save_settings()
            QMessageBox.information(self.ui, "Sucesso", "Parâmetros importados com sucesso!")
        except Exception as e:
            QMessageBox.critical(self.ui, "Erro", f"Erro ao importar: {str(e)}")



    # Exporta o gráfico atual de Curva PV como uma imagem
    def export_plot(self):
        import datetime
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        default_name = f"curva_pv_{timestamp}.png"
        filename, _ = QFileDialog.getSaveFileName(self.ui, "Exportar Gráfico", default_name, "Images (*.png *.jpg *.jpeg)")
        if not filename:
            return

        try:
            self.ui.pv_plot.export_plot(filename)
            QMessageBox.information(self.ui, "Sucesso", "Gráfico exportado com sucesso!")
        except Exception as e:
            QMessageBox.critical(self.ui, "Erro", f"Erro ao exportar: {str(e)}")

    # Exporta os resultados da simulação atual para um arquivo Excel
    def export_results(self):
        import datetime
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        default_name = f"resultados_{timestamp}.xlsx"
        filename, _ = QFileDialog.getSaveFileName(self.ui, "Exportar Resultados", default_name, "Excel Files (*.xlsx)")
        if not filename:
            return

        try:
            wb = Workbook()
            ws = wb.active

            # Exportar Resultados das Barras
            ws.append(["[Fluxo de Potência (Tensões Nodais)]"])
            headers = []
            for j in range(self.ui.table_bus_results.columnCount()):
                headers.append(self.ui.table_bus_results.horizontalHeaderItem(j).text())
            ws.append(headers)

            for i in range(self.ui.table_bus_results.rowCount()):
                row_data = []
                for j in range(self.ui.table_bus_results.columnCount()):
                    item = self.ui.table_bus_results.item(i, j)
                    row_data.append(item.text() if item else "")
                ws.append(row_data)
            ws.append([])

            # Exportar Resultados das Linhas
            ws.append(["[Fluxo nas Linhas]"])
            headers = []
            for j in range(self.ui.table_line_results.columnCount()):
                headers.append(self.ui.table_line_results.horizontalHeaderItem(j).text())
            ws.append(headers)

            for i in range(self.ui.table_line_results.rowCount()):
                row_data = []
                for j in range(self.ui.table_line_results.columnCount()):
                    item = self.ui.table_line_results.item(i, j)
                    row_data.append(item.text() if item else "")
                ws.append(row_data)

            wb.save(filename)
            QMessageBox.information(self.ui, "Sucesso", "Resultados exportados com sucesso!")
        except Exception as e:
            QMessageBox.critical(self.ui, "Erro", f"Erro ao exportar: {str(e)}")

    # Preenche o combobox de barras com barras do tipo 'pq' para a curva PV
    def populate_target_combo(self):
        self.ui.combo_target_bus.clear()
        for bus in self.state.buses.values():
            if bus.type == 'pq':
                self.ui.combo_target_bus.addItem(bus.name)

    # Chama a interface para desenhar a rede baseada no estado atual do sistema
    def update_diagram(self):
        self.ui.diagram_view.draw_network(self.state)

    def export_image(self):
        file_name, _ = QFileDialog.getSaveFileName(self.ui, "Salvar Imagem do Modelo", "diagrama_resultados.png", "Imagens PNG (*.png)")
        if file_name:
            self.ui.diagram_view.export_to_png(file_name)
            self.ui.toast.show_toast("Imagem exportada com sucesso!", True, self.ui)

    # Função que dispara a thread de simulação principal
    def run_simulation(self):
        self.ui.btn_export_image.setVisible(False)
        self.ui.btn_simulate.setEnabled(False)
        self.ui.btn_simulate.setText("Simulando...")
        self.ui.btn_simulate.setStyleSheet("background-color: #555555; color: #aaaaaa;")
        self.ui.progress_bar.setVisible(True)

        self.thread = SimulationThread(self.state)
        self.thread.finished.connect(self.on_simulation_finished)
        self.thread.start()

    # Callback invocado quando a thread de simulação termina sua execução
    def on_simulation_finished(self, results):
        self.ui.btn_simulate.setEnabled(True)
        self.ui.btn_simulate.setText("Iniciar Simulação")
        self.ui.btn_simulate.setStyleSheet("")
        self.ui.progress_bar.setVisible(False)

        if not results.success:
            self.ui.toast.show_toast("Erro: A simulação falhou.", False, self.ui)
            return

        self.ui.toast.show_toast("Simulação concluída com sucesso!", True, self.ui)

        # Atualizar Tabelas da UI
        bus_headers = ["Barra", "V (PU)", "Ângulo (°)", "P (MW)", "Q (MVAr)"]
        populate_table(self.ui.table_bus_results, results.bus_results, bus_headers)

        # Atualizar Cartões Flutuantes
        self.ui.diagram_view.update_results_cards(results.bus_results)
        self.ui.btn_export_image.setVisible(True)

        line_headers = ["Linha", "P_in (MW)", "Q_in (MVAr)", "P_out (MW)", "Q_out (MVAr)", "Perda (MW)", "Carga (%)"]
        populate_table(self.ui.table_line_results, results.line_results, line_headers)
        
        self.adjust_table_size(self.ui.table_bus_results)
        self.adjust_table_size(self.ui.table_line_results)

    # Função que dispara a thread de simulação da Curva PV
    def run_pv_simulation(self):
        self.ui.btn_simulate_pv.setEnabled(False)
        self.ui.btn_simulate_pv.setText("Gerando...")
        self.ui.btn_simulate_pv.setStyleSheet("background-color: #555555; color: #aaaaaa;")
        self.ui.progress_bar_pv.setVisible(True)

        target_bus = self.ui.combo_target_bus.currentText()

        self.pv_thread = PVCurveThread(self.state, target_bus)
        self.pv_thread.finished.connect(self.on_pv_simulation_finished)
        self.pv_thread.start()

    # Callback invocado quando a simulação de curva PV termina sua execução
    def on_pv_simulation_finished(self, results):
        self.ui.btn_simulate_pv.setEnabled(True)
        self.ui.btn_simulate_pv.setText("Gerar Curva PV")
        self.ui.btn_simulate_pv.setStyleSheet("")
        self.ui.progress_bar_pv.setVisible(False)

        if not results.success:
            self.ui.toast.show_toast("Erro: A geração da curva PV falhou.", False, self.ui)
            return

        self.ui.toast.show_toast("Curva PV gerada com sucesso!", True, self.ui)

        # Atualizar Gráfico
        target_bus = self.ui.combo_target_bus.currentText()
        self.ui.pv_plot.plot_curve(results.pv_curve_p, results.pv_curve_v, target_bus)

    # Inicializa e exibe a interface principal da aplicação
    def run(self):
        self.ui.show()
        sys.exit(self.app.exec())

if __name__ == '__main__':
    controller = MainController()
    controller.run()
